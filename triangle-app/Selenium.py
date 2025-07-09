from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import os
import json # For handling JSON from Gemini API
import requests # For making HTTP requests to Gemini API
from dotenv import load_dotenv
load_dotenv()

# Define the Excel report file name
EXCEL_REPORT_FILE = "BlackBox_testResult.xlsx"

def initialize_excel_report(filename):
    """
    Initializes the Excel report file with headers if it doesn't exist.
    """
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Triangle Test Results"
        # Write headers to the first row
        sheet.append(["X", "Y", "Z", "Actual Output (Raw)", "Expected Output", "Status"])
        workbook.save(filename)
        print(f"Created new Excel report: {filename}")
    else:
        print(f"Using existing Excel report: {filename}")

def read_test_cases_from_excel(filename):
    """
    Reads test cases (X, Y, Z, Expected Output) from the existing Excel file.
    Returns a list of tuples: (x, y, z, expected_type).
    """
    test_cases_from_excel = []
    if os.path.exists(filename):
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            # Iterate through rows starting from the second row (index 2) to skip headers
            for row_idx in range(2, sheet.max_row + 1):
                try:
                    x = float(sheet.cell(row=row_idx, column=1).value)
                    y = float(sheet.cell(row=row_idx, column=2).value)
                    z = float(sheet.cell(row=row_idx, column=3).value)
                    expected_type = str(sheet.cell(row=row_idx, column=5).value) # Column 5 for Expected Output
                    test_cases_from_excel.append((x, y, z, expected_type))
                except (ValueError, TypeError) as e:
                    print(f"Skipping row {row_idx} in Excel due to data error: {e}. Row data: {sheet.cell(row=row_idx, column=1).value}, {sheet.cell(row=row_idx, column=2).value}, {sheet.cell(row=row_idx, column=3).value}, {sheet.cell(row=row_idx, column=5).value}")
        except Exception as e:
            print(f"Error reading Excel file {filename}: {e}")
    return test_cases_from_excel

def add_or_update_test_result_to_excel(filename, x, y, z, actual_raw, expected, status_symbol):
    """
    Adds a new row or updates an existing row with test results to the Excel report file.
    It checks for an existing row with matching X, Y, Z values and updates it.
    If no match is found, a new row is appended.
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        
        found_row = False
        # Iterate through rows starting from the second row (index 2) to skip headers
        for row_idx in range(2, sheet.max_row + 1):
            # Read X, Y, Z values from the current row in the Excel sheet
            excel_x = sheet.cell(row=row_idx, column=1).value
            excel_y = sheet.cell(row=row_idx, column=2).value
            excel_z = sheet.cell(row=row_idx, column=3).value

            # Convert to float for comparison, handle potential None or non-numeric values
            try:
                excel_x_float = float(excel_x) if excel_x is not None else None
                excel_y_float = float(excel_y) if excel_y is not None else None
                excel_z_float = float(excel_z) if excel_z is not None else None
            except ValueError:
                # If conversion fails (e.g., cell contains non-numeric text), skip this row for comparison
                continue

            # Compare current test case inputs with Excel row inputs
            # Using a small tolerance for float comparisons might be necessary if values are generated with high precision
            # Here, direct comparison is used for simplicity assuming exact match needed.
            # You might consider using `math.isclose()` for float comparisons if precision issues arise
            if (excel_x_float == x and 
                excel_y_float == y and 
                excel_z_float == z):
                
                # Update the existing row with the new results
                sheet.cell(row=row_idx, column=4, value=actual_raw) # Actual Output (Raw)
                sheet.cell(row=row_idx, column=5, value=expected)   # Expected Output
                sheet.cell(row=row_idx, column=6, value=status_symbol) # Status (tick/cross)
                found_row = True
                print(f"Updated existing result in Excel: X={x}, Y={y}, Z={z}, Status: {status_symbol}")
                break # Exit loop once the matching row is updated

        if not found_row:
            # If no matching row was found after checking all existing rows, append a new row
            sheet.append([x, y, z, actual_raw, expected, status_symbol])
            print(f"Added new result to Excel: X={x}, Y={y}, Z={z}, Status: {status_symbol}")
            
        workbook.save(filename) # Save the workbook after changes

    except Exception as e:
        print(f"Error writing to Excel file {filename}: {e}")

def parse_actual_output_text(actual_webpage_text):
    """
    Parses the raw text output from the webpage to a standardized triangle type.
    This helps in robust comparison regardless of the exact phrasing on the page.
    """
    text_lower = actual_webpage_text.lower()
    if "equilateral" in text_lower:
        return "Equilateral"
    elif "isosceles" in text_lower:
        return "Isosceles"
    elif "scalene" in text_lower:
        return "Scalene"
    # Handle various "Not a triangle" messages, including invalid inputs
    elif "not a triangle" in text_lower or \
         "cannot form a triangle" in text_lower or \
         "invalid" in text_lower or \
         "positive numbers" in text_lower or \
         "sides must be greater than zero" in text_lower: # Added more specific invalid input checks
        return "Not a triangle"
    else:
        # Fallback for unexpected or new messages from the app
        return "Unknown or unexpected output"

def generate_edge_cases_with_gemini():
    """
    Generates additional triangle test cases, focusing on edge cases,
    using the Gemini API.
    """
    print("\nAttempting to generate additional edge cases using Gemini API...")
    generated_test_cases = []
    prompt = """
    Generate a diverse list of triangle side lengths (x, y, z) and their expected triangle type.
    Focus on edge cases and boundary conditions for triangle classification including Extreme values, Empty or null inputs, Invalid characters, Unusual user actions. 
    Include valid triangles (Equilateral, Isosceles, Scalene) and invalid triangles ("Not a triangle").
    For invalid triangles, include cases like:
    - Zero side length
    - Negative side length
    - Sum of two sides less than the third side
    - Sum of two sides exactly equal to the third side (degenerate triangle)
    - Very small positive numbers
    - Very large numbers
    - Decimal numbers that are very close to forming a specific type (e.g., almost equilateral due to float precision)
    

    Provide the output as a JSON array where each element is an object with 'x', 'y', 'z' (numbers), and 'expected_type' (string).
    Example: [{"x": 3, "y": 4, "z": 5, "expected_type": "Scalene"}]
    """
    chatHistory = []
    chatHistory.append({"role": "user", "parts": [{"text": prompt}]})

    # The API key should be provided by the environment, leaving it empty as per instructions.
    apiKey = os.getenv("GOOGLE_API_KEY")
    apiUrl = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={apiKey}"

    payload = {
        "contents": chatHistory,
        "generationConfig": {
            "responseMimeType": "application/json",
            "responseSchema": {
                "type": "ARRAY",
                "items": {
                    "type": "OBJECT",
                    "properties": {
                        "x": {"type": "NUMBER"},
                        "y": {"type": "NUMBER"},
                        "z": {"type": "NUMBER"},
                        "expected_type": {"type": "STRING"}
                    },
                    "required": ["x", "y", "z", "expected_type"]
                }
            }
        }
    }

    try:
        response = requests.post(apiUrl, headers={'Content-Type': 'application/json'}, json=payload)
        response.raise_for_status() # Raise an exception for HTTP errors (4xx or 5xx)
        result = response.json()

        if result.get("candidates") and result["candidates"][0].get("content") and \
           result["candidates"][0]["content"].get("parts") and \
           result["candidates"][0]["content"]["parts"][0].get("text"):
            
            json_string = result["candidates"][0]["content"]["parts"][0]["text"]
            # The API might return JSON as a string that needs to be parsed
            generated_data = json.loads(json_string)
            
            # Validate the structure of generated_data
            if isinstance(generated_data, list):
                for case in generated_data:
                    # Ensure all required keys exist and types are correct for comparison
                    if (isinstance(case, dict) and 
                        isinstance(case.get("x"), (int, float)) and
                        isinstance(case.get("y"), (int, float)) and
                        isinstance(case.get("z"), (int, float)) and
                        isinstance(case.get("expected_type"), str)):
                        generated_test_cases.append(
                            (case["x"], case["y"], case["z"], case["expected_type"])
                        )
                    else:
                        print(f"Skipping malformed generated test case: {case}")
            print(f"Generated {len(generated_test_cases)} additional edge cases from Gemini API.")
        else:
            print("Gemini API response did not contain expected content or was empty.")
            print(f"Full API response: {json.dumps(result, indent=2)}")

    except requests.exceptions.RequestException as req_err:
        print(f"Network or API request error: {req_err}")
    except json.JSONDecodeError as json_err:
        print(f"Failed to decode JSON from Gemini API response: {json_err}")
        if response and response.text:
            print(f"Raw API response text: {response.text}")
    except Exception as e:
        print(f"An unexpected error occurred during Gemini API call: {e}")
    
    return generated_test_cases

def test_triangle_app(x, y, z, expected_type):
    """
    Tests the Triangle App with given inputs (x, y, z) and verifies the output
    against the expected triangle type. Records the result in an Excel file.
    """
    driver = None
    actual_raw_output = "N/A" # Stores the exact text from the webpage
    parsed_actual_type = "N/A" # Stores the standardized parsed type
    test_status_symbol = "✗" # Default to cross (fail) symbol

    try:
        # Initialize the Chrome driver. Ensure chromedriver is in your PATH.
        driver = webdriver.Chrome()
        driver.get("https://vekinfortesting.github.io/triangle-app/")

        wait = WebDriverWait(driver, 10) # Wait up to 10 seconds for elements

        # Locate input fields using their IDs as specified ('side1', 'side2', 'side3')
        x_input = wait.until(EC.visibility_of_element_located((By.ID, "side1")))
        y_input = driver.find_element(By.ID, "side2")
        z_input = driver.find_element(By.ID, "side3")
        check_button = driver.find_element(By.TAG_NAME, "button") # Locate the button

        # Clear any pre-filled values in the input fields
        x_input.clear()
        y_input.clear()
        z_input.clear()

        # Send the test data to the input fields
        x_input.send_keys(str(x))
        y_input.send_keys(str(y))
        z_input.send_keys(str(z))

        # Click the 'Check Triangle' button
        check_button.click()

        # Wait for the result element to become visible and get its text
        result_element = wait.until(EC.visibility_of_element_located((By.ID, "result")))
        actual_raw_output = result_element.text # Get the raw text from the webpage

        # Parse the raw output to a standardized type for comparison
        parsed_actual_type = parse_actual_output_text(actual_raw_output)

        print(f"Testing X={x}, Y={y}, Z={z}")
        print(f"  Actual (Raw): '{actual_raw_output}'")
        print(f"  Expected: '{expected_type}'")
        print(f"  Parsed Actual Type: '{parsed_actual_type}'")

        # Compare the parsed actual type with the expected type
        if parsed_actual_type == expected_type:
            test_status_symbol = "✓" # Tick symbol for pass
            print(f"  Test passed for X={x}, Y={y}, Z={z}\n")
        else:
            test_status_symbol = "✗" # Cross symbol for fail
            print(f"  Test FAILED for X={x}, Y={y}, Z={z}. Expected '{expected_type}' but got '{parsed_actual_type}'\n")

    except Exception as e:
        # Catch any exception during test execution (e.g., element not found, network issues)
        print(f"  An error occurred during testing X={x}, Y={y}, Z={z}: {e}")
        actual_raw_output = f"ERROR: {e}" # Log the error message in the actual output column
        parsed_actual_type = "ERROR" # Indicate an error in the parsed type
        test_status_symbol = "✗" # An error in test execution means a failed test case

    finally:
        # Add or update the test result in the Excel file regardless of success or failure
        add_or_update_test_result_to_excel(EXCEL_REPORT_FILE, x, y, z, actual_raw_output, expected_type, test_status_symbol)
        if driver:
            driver.quit() # Close the browser for the current test case

# --- Main execution block for test cases ---
if __name__ == "__main__":
    print("--- Starting Triangle App Automated Tests ---\n")

    # Initialize the Excel report file (create if not exists, add headers)
    initialize_excel_report(EXCEL_REPORT_FILE)

    # Phase 1: Read and re-test existing cases from Excel
    excel_test_cases = read_test_cases_from_excel(EXCEL_REPORT_FILE)
    if excel_test_cases:
        print(f"--- Running {len(excel_test_cases)} test cases from '{EXCEL_REPORT_FILE}' ---")
        for x, y, z, expected in excel_test_cases:
            test_triangle_app(x, y, z, expected)
            time.sleep(0.5) # Short delay between tests for stability
    else:
        print(f"No existing test cases found in '{EXCEL_REPORT_FILE}'.")

    # Phase 2: Generate and test new edge cases using Gemini API
    gemini_generated_cases = generate_edge_cases_with_gemini()
    if gemini_generated_cases:
        print(f"--- Running {len(gemini_generated_cases)} AI-generated edge cases ---")
        for x, y, z, expected in gemini_generated_cases:
            test_triangle_app(x, y, z, expected)
            time.sleep(0.5) # Short delay between tests for stability
    else:
        print("No new edge cases were generated by Gemini API.")

    print("\n--- All Triangle App Tests Completed. Check 'test_results.xlsx' for detailed results. ---")
