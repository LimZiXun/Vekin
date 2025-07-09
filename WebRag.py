from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import os
import json # For handling JSON from Gemini API
import requests # For making HTTP requests to Gemini API
from datetime import datetime # For timestamp in log file
from dotenv import load_dotenv
load_dotenv()

# Define the Excel report file names
EXCEL_REPORT_FILE = "test_results.xlsx"
TEST_LOG_FILE = "test_run_log.xlsx" # New log file

def initialize_excel_report(filename):
    """
    Initializes the main Excel report file with headers if it doesn't exist.
    """
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Triangle Test Results"
        # Write headers to the first row
        sheet.append(["X", "Y", "Z", "Actual Output (Raw)", "Expected Output", "Status"])
        workbook.save(filename)
        print(f"Created new Excel report: {filename}")
    else:
        print(f"Using existing Excel report: {filename}")

def initialize_log_report(filename):
    """
    Initializes the Excel log file with headers if it doesn't exist.
    """
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Test Run Logs"
        # Write headers to the first row
        sheet.append(["Date & Time", "Total Test Cases Run", "Passed", "Failed", "Errors"])
        workbook.save(filename)
        print(f"Created new test log file: {filename}")
    else:
        print(f"Using existing test log file: {filename}")

def read_test_cases_from_excel(filename):
    """
    Reads test cases (X, Y, Z, Expected Output) from the existing Excel file.
    Returns a list of tuples: (x, y, z, expected_type).
    """
    test_cases_from_excel = []
    if os.path.exists(filename):
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            # Iterate through rows starting from the second row (index 2) to skip headers
            for row_idx in range(2, sheet.max_row + 1):
                try:
                    x = float(sheet.cell(row=row_idx, column=1).value)
                    y = float(sheet.cell(row=row_idx, column=2).value)
                    z = float(sheet.cell(row=row_idx, column=3).value)
                    expected_type = str(sheet.cell(row=row_idx, column=5).value) # Column 5 for Expected Output
                    test_cases_from_excel.append((x, y, z, expected_type))
                except (ValueError, TypeError) as e:
                    print(f"Skipping row {row_idx} in Excel due to data error: {e}. Row data: {sheet.cell(row=row_idx, column=1).value}, {sheet.cell(row=row_idx, column=2).value}, {sheet.cell(row=row_idx, column=3).value}, {sheet.cell(row=row_idx, column=5).value}")
        except Exception as e:
            print(f"Error reading Excel file {filename}: {e}")
    return test_cases_from_excel

def add_or_update_test_result_to_excel(filename, x, y, z, actual_raw, expected, status_symbol):
    """
    Adds a new row or updates an existing row with test results to the Excel report file.
    It checks for an existing row with matching X, Y, Z values and updates it.
    If no match is found, a new row is appended.
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        
        found_row = False
        # Iterate through rows starting from the second row (index 2) to skip headers
        for row_idx in range(2, sheet.max_row + 1):
            # Read X, Y, Z values from the current row in the Excel sheet
            excel_x = sheet.cell(row=row_idx, column=1).value
            excel_y = sheet.cell(row=row_idx, column=2).value
            excel_z = sheet.cell(row=row_idx, column=3).value

            # Convert to float for comparison, handle potential None or non-numeric values
            try:
                excel_x_float = float(excel_x) if excel_x is not None else None
                excel_y_float = float(excel_y) if excel_y is not None else None
                excel_z_float = float(excel_z) if excel_z is not None else None
            except ValueError:
                # If conversion fails (e.g., cell contains non-numeric text), skip this row for comparison
                continue

            # Compare current test case inputs with Excel row inputs
            if (excel_x_float == x and 
                excel_y_float == y and 
                excel_z_float == z):
                
                # Update the existing row with the new results
                sheet.cell(row=row_idx, column=4, value=actual_raw) # Actual Output (Raw)
                sheet.cell(row=row_idx, column=5, value=expected)   # Expected Output
                sheet.cell(row=row_idx, column=6, value=status_symbol) # Status (tick/cross)
                found_row = True
                # print(f"Updated existing result in Excel: X={x}, Y={y}, Z={z}, Status: {status_symbol}") # Suppress for less console spam
                break # Exit loop once the matching row is updated

        if not found_row:
            # If no matching row was found after checking all existing rows, append a new row
            sheet.append([x, y, z, actual_raw, expected, status_symbol])
            # print(f"Added new result to Excel: X={x}, Y={y}, Z={z}, Status: {status_symbol}") # Suppress for less console spam
            
        workbook.save(filename) # Save the workbook after changes

    except Exception as e:
        print(f"Error writing to Excel file {filename}: {e}")

def add_log_entry(filename, total_cases, passed_cases, failed_cases, error_cases):
    """
    Appends a new log entry to the test run log file.
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([timestamp, total_cases, passed_cases, failed_cases, error_cases])
        workbook.save(filename)
        print(f"\nTest run summary logged to '{filename}':")
        print(f"  Total: {total_cases}, Passed: {passed_cases}, Failed: {failed_cases}, Errors: {error_cases}")
    except Exception as e:
        print(f"Error writing to log file {filename}: {e}")

def parse_actual_output_text(actual_webpage_text):
    """
    Parses the raw text output from the webpage to a standardized triangle type.
    This helps in robust comparison regardless of the exact phrasing on the page.
    """
    text_lower = actual_webpage_text.lower()
    if "equilateral" in text_lower:
        return "Equilateral"
    elif "isosceles" in text_lower:
        return "Isosceles"
    elif "scalene" in text_lower:
        return "Scalene"
    # Handle various "Not a triangle" messages, including invalid inputs
    elif "not a triangle" in text_lower or \
         "cannot form a triangle" in text_lower or \
         "invalid" in text_lower or \
         "positive numbers" in text_lower or \
         "sides must be greater than zero" in text_lower:
        return "Not a triangle"
    else:
        # Fallback for unexpected or new messages from the app
        return "Unknown or unexpected output"

def generate_edge_cases_with_gemini():
    """
    Generates additional triangle test cases, focusing on edge cases,
    using the Gemini API.
    """
    print("\nAttempting to generate additional edge cases using Gemini API...")
    generated_test_cases = []
    prompt = """
    Generate a diverse list of triangle side lengths (x, y, z) and their expected triangle type.
    Focus on edge cases and boundary conditions for triangle classification.
    Include valid triangles (Equilateral, Isosceles, Scalene) and invalid triangles ("Not a triangle").
    For invalid triangles, include cases like:
    - Zero side length
    - Negative side length
    - Sum of two sides less than the third side
    - Sum of two sides exactly equal to the third side (degenerate triangle)
    - Very small positive numbers
    - Very large numbers
    - Decimal numbers that are very close to forming a specific type (e.g., almost equilateral due to float precision)
    - Edge cases like (0.0001, 0.0001, 0.0001) for equilateral
    - Extreme large numbers (1000000, 1000000, 1000000)

    Provide the output as a JSON array where each element is an object with 'x', 'y', 'z' (numbers), and 'expected_type' (string).
    Example: [{"x": 3, "y": 4, "z": 5, "expected_type": "Scalene"}]
    """
    chatHistory = []
    chatHistory.append({"role": "user", "parts": [{"text": prompt}]})

    # The API key should be loaded from environment variables (e.g., .env file)
    apiKey = ""
    if not apiKey:
        print("Error: GOOGLE_API_KEY environment variable not set. Cannot call Gemini API.")
        return []

    apiUrl = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={apiKey}"

    payload = {
        "contents": chatHistory,
        "generationConfig": {
            "responseMimeType": "application/json",
            "responseSchema": {
                "type": "ARRAY",
                "items": {
                    "type": "OBJECT",
                    "properties": {
                        "x": {"type": "NUMBER"},
                        "y": {"type": "NUMBER"},
                        "z": {"type": "NUMBER"},
                        "expected_type": {"type": "STRING"}
                    },
                    "required": ["x", "y", "z", "expected_type"]
                }
            }
        }
    }

    try:
        response = requests.post(apiUrl, headers={'Content-Type': 'application/json'}, json=payload)
        response.raise_for_status() # Raise an exception for HTTP errors (4xx or 5xx)
        result = response.json()

        if result.get("candidates") and result["candidates"][0].get("content") and \
           result["candidates"][0]["content"].get("parts") and \
           result["candidates"][0]["content"]["parts"][0].get("text"):
            
            json_string = result["candidates"][0]["content"]["parts"][0]["text"]
            generated_data = json.loads(json_string)
            
            if isinstance(generated_data, list):
                for case in generated_data:
                    if (isinstance(case, dict) and 
                        isinstance(case.get("x"), (int, float)) and
                        isinstance(case.get("y"), (int, float)) and
                        isinstance(case.get("z"), (int, float)) and
                        isinstance(case.get("expected_type"), str)):
                        generated_test_cases.append(
                            (case["x"], case["y"], case["z"], case["expected_type"])
                        )
                    else:
                        print(f"Skipping malformed generated test case: {case}")
            print(f"Generated {len(generated_test_cases)} additional edge cases from Gemini API.")
        else:
            print("Gemini API response did not contain expected content or was empty.")
            print(f"Full API response: {json.dumps(result, indent=2)}")

    except requests.exceptions.RequestException as req_err:
        print(f"Network or API request error: {req_err}")
    except json.JSONDecodeError as json_err:
        print(f"Failed to decode JSON from Gemini API response: {json_err}")
        if response and response.text:
            print(f"Raw API response text: {response.text}")
    except Exception as e:
        print(f"An unexpected error occurred during Gemini API call: {e}")
    
    return generated_test_cases

def test_triangle_app(x, y, z, expected_type):
    """
    Tests the Triangle App with given inputs (x, y, z) and verifies the output
    against the expected triangle type. Records the result in an Excel file.
    Returns the status ('PASS', 'FAIL', 'ERROR').
    """
    driver = None
    actual_raw_output = "N/A"
    parsed_actual_type = "N/A"
    test_status_symbol = "✗" # Default to cross (fail) symbol
    result_status = "FAIL" # Default status for return value

    try:
        driver = webdriver.Chrome()
        driver.get("https://vekinfortesting.github.io/triangle-app/")

        wait = WebDriverWait(driver, 10)

        x_input = wait.until(EC.visibility_of_element_located((By.ID, "side1")))
        y_input = driver.find_element(By.ID, "side2")
        z_input = driver.find_element(By.ID, "side3")
        check_button = driver.find_element(By.TAG_NAME, "button")

        x_input.clear()
        y_input.clear()
        z_input.clear()

        x_input.send_keys(str(x))
        y_input.send_keys(str(y))
        z_input.send_keys(str(z))

        check_button.click()

        result_element = wait.until(EC.visibility_of_element_located((By.ID, "result")))
        actual_raw_output = result_element.text
        parsed_actual_type = parse_actual_output_text(actual_raw_output)

        print(f"Testing X={x}, Y={y}, Z={z}")
        print(f"  Actual (Raw): '{actual_raw_output}'")
        print(f"  Expected: '{expected_type}'")
        print(f"  Parsed Actual Type: '{parsed_actual_type}'")

        if parsed_actual_type == expected_type:
            test_status_symbol = "✓" # Tick symbol
            result_status = "PASS"
            print(f"  Test passed for X={x}, Y={y}, Z={z}\n")
        else:
            test_status_symbol = "✗" # Cross symbol
            result_status = "FAIL"
            print(f"  Test FAILED for X={x}, Y={y}, Z={z}. Expected '{expected_type}' but got '{parsed_actual_type}'\n")

    except Exception as e:
        print(f"  An error occurred during testing X={x}, Y={y}, Z={z}: {e}")
        actual_raw_output = f"ERROR: {e}"
        parsed_actual_type = "ERROR"
        test_status_symbol = "✗" # Cross symbol for error
        result_status = "ERROR" # Indicate an error in test execution

    finally:
        add_or_update_test_result_to_excel(EXCEL_REPORT_FILE, x, y, z, actual_raw_output, expected_type, test_status_symbol)
        if driver:
            driver.quit()
        return result_status

# --- Main execution block for test cases ---
if __name__ == "__main__":
    print("--- Starting Triangle App Automated Tests ---\n")

    # Initialize both Excel report files
    initialize_excel_report(EXCEL_REPORT_FILE)
    initialize_log_report(TEST_LOG_FILE)

    total_tests_run = 0
    passed_tests = 0
    failed_tests = 0
    error_tests = 0

    # Phase 1: Read and re-test existing cases from Excel
    excel_test_cases = read_test_cases_from_excel(EXCEL_REPORT_FILE)
    if excel_test_cases:
        print(f"--- Running {len(excel_test_cases)} test cases from '{EXCEL_REPORT_FILE}' ---")
        for x, y, z, expected in excel_test_cases:
            total_tests_run += 1
            status = test_triangle_app(x, y, z, expected)
            if status == "PASS":
                passed_tests += 1
            elif status == "FAIL":
                failed_tests += 1
            else: # "ERROR"
                error_tests += 1
            time.sleep(0.5) # Short delay between tests for stability
    else:
        print(f"No existing test cases found in '{EXCEL_REPORT_FILE}'.")

    # Phase 2: Generate and test new edge cases using Gemini API
    gemini_generated_cases = generate_edge_cases_with_gemini()
    if gemini_generated_cases:
        print(f"--- Running {len(gemini_generated_cases)} AI-generated edge cases ---")
        for x, y, z, expected in gemini_generated_cases:
            total_tests_run += 1
            status = test_triangle_app(x, y, z, expected)
            if status == "PASS":
                passed_tests += 1
            elif status == "FAIL":
                failed_tests += 1
            else: # "ERROR"
                error_tests += 1
            time.sleep(0.5) # Short delay between tests for stability
    else:
        print("No new edge cases were generated by Gemini API.")

    # Log the summary of the test run
    add_log_entry(TEST_LOG_FILE, total_tests_run, passed_tests, failed_tests, error_tests)

    print("\n--- All Triangle App Tests Completed. Check 'test_results.xlsx' for detailed results and 'test_run_log.xlsx' for run summary. ---")
